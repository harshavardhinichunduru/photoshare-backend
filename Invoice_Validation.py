from datetime import datetime
print("Start Time:", datetime.now().strftime("%H:%M:%S"))
import os, sys, time
import logging
import Process_Image as pi
import openpyxl

# hostname = "172.17.104.109" #production
hostname = "192.168.247.109" #after ivanti implemented
username = "root"
password = "Sonora2000!"
port = 22  # Standard SSH port


# Dictionary mapping market values to the corresponding text
market_lla_map = {
    "SXM": "UNITED TELECOMMUNICATION SERVICES SINT MAARTEN N.V.",
    "CUR": "Antelecom NV",
    "BON": "United Telecommunication Services N.V.",
    "SABA": "WINDWARD ISLANDS CELLULAR COMPANY NV",
    "EUX" : "WINDWARD ISLANDS CELLULAR COMPANY NV"
}

# 	   	      MKT_CODE|       LANGUAGE_CODE|SHORT_DISPLAY  |DISPLAY_VALUE            
# --------------------|--------------------|---------------|-------------------------
#                    1|                   1|CUR            |Curacao                  
#                    2|                   1|BON            |Bonaire                  
#                   10|                   1|SMF            |St.Martin (French)       
#                   11|                   1|SAB            |Saba                     
#                   12|                   1|EUX            |St. Eustatius            
#                    7|                   1|KNA            |St. Kitts & Nevis        
#                    8|                   1|SUR            |Suriname                 
#                    9|                   1|SXM            |St.Maarten (Dutch)       
#                   17|                   1|SBH            |St.Barths                
#                   15|                   1|CUREZ          |Curacao E-Zone           
#                   16|                   1|SXMEZ          |St.Maarten (Dutch) E-Zone

# Dictionary mapping LLA Name values to the corresponding text
lla_name = {
    "10":"UTS CARAIBE SARL",
    "7":"UTS CARIGLOBE MOBILE SERVICE LTD",
    "11":"WINDWARD ISLANDS CELLULAR COMPANY NV",
    "12":"WINDWARD ISLANDS CELLULAR COMPANY NV",
    "9":"UNITED TELECOMMUNICATION SERVICES SINT MAARTEN N.V.",
    "16":"UNITED TELECOMMUNICATION SERVICES SINT MAARTEN N.V.",
    "20":"Kelcom International Antigua and Barbuda Limited.",
    "2":"United Telecommunication Services N.V.",
    "1":"Berg Arrarat 1, Willemstad, Curaçao"
}

lla_address = {
    "10":"24 rue de la République, Marigot, 97150 Saint-Martin",
    "7":"P.O.BOX 2013, BASTERRE, ST.KITTS",
    "11":"P.O. BOX 1, THE BOTTOM, SABA",
    "12":"MAZINGA MALL, FORT ORANJESTAD, ST. EUSTATIUS",
    "9":"Codville Webster Road 2, Philipsburg, St. Maarten",
    "16":"Codville Webster Road 2, Philipsburg, St. Maarten",
    "20":{
        "2": "ABST ID # 0178455", 
        "3" : "P.O. Box 65"
        },
    "2":"Kaya Grandi 32, Kralendijk, Bonaire",
    "1":"Berg Arrarat 1, Willemstad, Curaçao",
    "15":"Berg Arrarat 1, Willemstad, Curaçao"
}

def is_file_available(file_path):
    if os.path.exists(file_path) or os.path.isfile(file_path):
        return True 
    else:
        log_message(f'Error: File {file_path} does not exist',True)
        return False
    
PointCompletion = {}
tbl_account_details ={}
tbl_bill_details ={}
tbl_cmf_details ={}
threads = {}
Excel_Input=[]
Proposal_text_list ={}
no_rows_account=[]

Is_success = True

def setup_logging(log_file):
    logging.basicConfig(filename=log_file, level=logging.INFO, format='%(asctime)s:%(levelname)s:%(message)s')

def log_message(message,Error=False):
    if Error:
        global Is_success
        Is_success=False
    logging.info(message)
    print(message)

def get_lla_address(mkt_code,owning_cost_ctr):
    value = lla_address[mkt_code]
    if isinstance(value, dict):
        return value[owning_cost_ctr]
    else:
        return value
    
def refresh():
    global PointCompletion
    PointCompletion = {
    "Point1" : True, 
    "Point2" : True,
    "Point3" : True,
    "Point4" : True,
    "Point5" : True,
    "Point6" : True,
    "Point7" : True,
    "Point8" : True,
    "Point9" : True,
    "Point10" : True,
    "Point11" : True,
    "Point12" : True,
    "Point13" : True,
    "Point14" : True,
    "Point15" : True,
    "Point16" : True,
    "Point17" : True,
    "Point18" : True,
    "Point19" : True,
    }
# Example usage:
setup_logging('Invoice_validation.log')

def connect():
    try:        
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname, port, username, password)
        return True

    except paramiko.AuthenticationException:
        log_message('Error: Authentication failed',True)
        return False
    except paramiko.SSHException as sshException:
        log_message(f'Error: An SSH error occurred: {sshException}',True)
        return False
    except Exception as e:
        log_message(f'Error: An unexpected error occurred: {e}',True)
        return False

def disconnect():
    if client:
        client.close()
        log_message("Connection closed.")

def execute(command):
    try:
        stdin, stdout, stderr = client.exec_command(command)
        output = stdout.read().decode()
        error = stderr.read().decode()

        # print(output)
        # result.append({command:output})
        if error:
            print("Command error:")
            log_message(f'{error}',True)
            # result.append({command:error})
        return output

    except paramiko.AuthenticationException:
        log_message("Error: Authentication failed",True)
    except paramiko.SSHException as sshException:
        log_message(f'"Error: An SSH error occurred:" {sshException}',True)
    except Exception as e:
        log_message(f'Error: An unexpected error occurred: {e}',True)

def writeScript(data_to_write):
    try:
        # Open the remote file for writing (will create if it doesn't exist)
        # Remote file path
        remote_file = "/tmp/tmp.sh"
        sftp = client.open_sftp()
        with sftp.open(remote_file, 'w') as remote_file_handle:
            # Write data to the remote file
            remote_file_handle.write(data_to_write.encode())  # Encode data to bytes
    except IOError as e:
        log_message(f"IOError: {e}")
    except Exception as e:
        log_message(f"Unable to connect for ftp {hostname}",True)
        log_message(f"An error occurred: {e}",True)
    finally:
        sftp.close()

    # print(f"Data written successfully to {remote_file}")

def collect_tbl_details():
    global tbl_account_details
    global tbl_cmf_details
    global tbl_bill_details
    global Excel_Input
    global Proposal_text_list
    global no_rows_account
    
    def make_query(query):
        content="""
#!/bin/bash
export ORACLE_BASE=/home/oracle/product
export ORACLE_HOME=/home/oracle/product/11.2.0/db_1
export PATH=${PATH}:$ORACLE_HOME/bin

sqlplus -s cbs_owner/comverse@cust1 << THEEND
SET SERVEROUTPUT ON FORMAT WRAPPED SIZE 1000000;
set head off;
set trimout on
set lines 32767
set numwidth 20
set pagesize 15000
set colsep '|'
WHENEVER SQLERROR EXIT FAILURE ROLLBACK;
WHENEVER OSERROR EXIT FAILURE ROLLBACK;

"""+ query +"""
quit
THEEND
"""
        return content
    
    def execute_sql(query):
        content = make_query(query)
        writeScript(content)
        output = execute('cd /tmp; chmod +x tmp.sh; ./tmp.sh')
        if output.find('no rows selected') > -1 or len(output) == 0:
            log_message('Error: No rows found',True)
            return []
        # Convert the output to a list of rows (assuming newline-separated rows)
        rows = output.split("\n")
        # Remove empty or whitespace-only items
        cleaned_rows = [row for row in rows if row.strip()]
        if len(cleaned_rows) == 1:
            return cleaned_rows
        return cleaned_rows
    
    def func1(account_no,STATEMENT_DATE):
        #get child's external_id list 
        query = f'''select external_id from external_id_equip_map_view where INACTIVE_DATE is null and account_no in (select account_no 
  from customer_id_acct_map where external_id in ('{account_no}'));'''
        
        external_id_list = execute_sql(query)
        if len(external_id_list) == 0:
            log_message(f'No rows found for the account {account_no}', True)
            no_rows_account.append(account_no)
            return {}
        
        query = f'''WITH external_id_list as (
  select external_id from external_id_equip_map_view where INACTIVE_DATE is null and account_no in (select account_no 
  from customer_id_acct_map where external_id in ('{account_no}'))
), account_list as (
select distinct b.external_id, a.account_no 
  from external_id_equip_map_view a, external_id_list b
  where 
  INACTIVE_DATE IS NULL 
  and a.external_id = b.external_id
), bill_list as (
select b.external_id, b.account_no, max(a.bill_ref_no) as bill_ref_no
  from bill_invoice a, account_list b 
  where 
  a.account_no = b.account_no 
  and a.prep_status = 4
  AND a.prep_error_code IS NULL
  AND a.statement_date = TO_DATE('{STATEMENT_DATE}', 'DD MON YY')
  group by b.external_id, b.account_no
)
select distinct b.external_id, c.PROPOSAL_TEXT
  from bill_invoice_detail a, bill_list b, vw_offers c
  where 
  a.bill_ref_no = b.bill_ref_no
  AND a.type_code IN (2,3)
  AND a.subscr_no > 0
  AND a.AMOUNT_POSTPAID > 0
  and rc_term_inst_id is not null
  and a.offer_id = c.offer_id ;
'''
#         queries = []
#         for external_id in external_id_list: 
#             queries.append(f'''select  distinct('{external_id}' || '|' || PROPOSAL_TEXT) from offer_values 
#   where LANGUAGE_CODE =1 and 	reseller_version_id in(SELECT max(reseller_version_id) FROM reseller_version) and 
#   offer_id in (
#     select offer_id from offer_inst@main1 where offer_inst_id in (
#       select offer_inst_id from rc_term_inst where rc_term_inst_id in (
#         select rc_term_inst_id from bill_invoice_detail where bill_ref_no in (
#           select  max(bill_ref_no) from bill_invoice where account_no in 
# 		    (select account_no from external_id_equip_map_view where external_id='{external_id}' and INACTIVE_DATE is null)
#           and prep_status=4 and prep_error_code is  null and statement_date = to_date('{STATEMENT_DATE}','DD MON YY') )
#         and type_code in (2,3)
#         and subscr_no > 0
#         and AMOUNT_POSTPAID>0)))
# ''')
# #         big_query = ''
# #         for i in range(len(query)-1):
# #             big_query += query[i]
# #             big_query += '''union all
# # '''
#         # Batch size (keep below 300 to be safe)
#         BATCH_SIZE = 250

#         rows=[]
#         batched_queries = [queries[i:i + BATCH_SIZE] for i in range(0, len(queries), BATCH_SIZE)]
#         for batch in batched_queries:
#             big_query = " UNION ALL ".join(batch)  # Merge queries in batch
#             big_query += ';'
#             rows.extend(execute_sql(big_query))
        
        rows = execute_sql(query)
        # big_query += query[-1] + ';'
        # rows =execute_sql(big_query)
        if len(rows) == 0:
            log_message(f'No child account found for the account {account_no}', True)
            no_rows_account.append(account_no)
            return {}
        # Convert rows into a dictionary
        data = {}
        for row in rows:
            row = ' '.join(row.split())
            columns = row.replace(' ','').split('|')
            
            if len(columns) == 2:
                EXTERNAL_ID,PROPOSAL_TEXT = columns
                if account_no not in data.keys():
                    data[account_no] =[]
                data[account_no].append({"EXTERNAL_ID":EXTERNAL_ID,"PROPOSAL_TEXT":PROPOSAL_TEXT})
            else:
                print(f"Skipping malformed row: {row}")
        return data
        
    external_id_list=[]
    if len(Excel_Input.keys()) == 1:
        external_id_list = list(Excel_Input.keys())
    if len(Excel_Input.keys()) > 1:
        external_id_list = list(Excel_Input.keys())

    if not connect():
        log_message(f'Connection failed while verify service for child account', True)
        return False
    
    #new changes 23-jan-2025
    account_list = []
    for account_number in external_id_list:
        account_list.append([account_number, Excel_Input[account_number]['Invoice_Date']])
    
        data = func1(account_number, Excel_Input[account_number]['Invoice_Date'])
        Proposal_text_list.update(data)
    
    disconnect()



def readExcel(excel_file):
    # Read the excel file and store data in a dictionary with key as Account Number & Value as DataFrame
    dfs = pd.read_excel(excel_file,converters={"BILL_REF_NO": str,"Account_Number": str})
    dic_entry={}
    from_excel = {}
    for row in range(len(dfs)):
        for column in dfs.columns:
            dic_entry[column] = dfs.loc[row,column]
        acc_num = str(dfs.loc[row,'Account_Number'])
        if acc_num in from_excel.keys():
            log_message("Error: Duplicated entry for Account_No [" + acc_num + "]",True)
        else:
            from_excel[acc_num] = dic_entry.copy()
        dic_entry={}
    return from_excel

class handlingPDF:
    def __init__(self):
        self.account_number = ''
        self.pdf_file = ''
        self.sample_pdf = ''
        self.profile = ''
        self.__converted_content =''
        self.__from_pdf={}
        self.__pages_out = []
        self.from_excel = {}
        self.__lang_string={}
        self.account_category = ''
        self.file_path=''
        # this (below) will have text messages in all languages.
        #sample: {'English':['Page','INVOICE','Account||number']}
        #sample: {'Dutch':['Página','INVOICE','Account||number']}
        self.__string_content={} 

    def read_xls_string(self,excel_file):
        lang=self.from_excel[self.account_number]['LANGUAGE'] 
        if not lang:
            log_message("Error: Language is None",True)
            sys.exit(0)
        #lang='English'
        self.__lang_string={}
        dfs = pd.read_excel(excel_file, sheet_name='Name')
        for row in range(len(dfs)):
            key=dfs.loc[row,'name_id']
            value=dfs.loc[row,lang]
            self.__lang_string[key]= value
        
    def order_content(self,content):
        data = []
        for line in content:
            values=line.split(',')
            if len(values) > 2:
                tmp={"x_position": float(values[0].replace('(','')), "y_position": float(values[1]), "message": line.split(',', 2)[-1]}
                data.append(tmp)

        # Sort data by y_position ascending, then by x_position descending
        sorted_data = sorted(data, key=lambda x: (-x["y_position"], x["x_position"]))
        output=[]

        for item in sorted_data:
            x_pos = str(item['x_position'])
            y_pos = str(item['y_position'])
            msg = item['message']
            line='(' + x_pos + ',' + y_pos + ',' + msg 
            output.append(line)

        return output

    def extract_text_and_coordinates(self,pdf_file):
        fp = open(pdf_file, 'rb')
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        pages = PDFPage.get_pages(fp)

        output = []
        final_out = []
        for page in pages:
            # print('Info: Processing next page...')
            interpreter.process_page(page)
            layout = device.get_result()
            for lobj in layout:
                if isinstance(lobj, LTTextBox):
                    x, y, text = lobj.bbox[0], lobj.bbox[3], lobj.get_text()
                    output.append(f"({x},{y},{text})")
            converted_txt = self.convert_content('\n'.join(output))
            ordered_txt = self.order_content(converted_txt)
            self.__pages_out.append('\n'.join(ordered_txt))
            for line in ordered_txt:
                final_out.append(line)
            output=[]
        return '\n'.join(final_out)

    # Function to convert the content
    def convert_content(self,input_text):
        # Split the input text into lines
        lines = input_text.strip().split('\n')
        # Process each line
        processed_lines = []
        tmp_lines=[]
        tmp=""
        for line in lines:
            # Remove any trailing whitespace and newlines
            clean_line = line.strip()
            if clean_line.find(')') != -1:  # If there'
                tmp_lines.append(clean_line)
                for l in tmp_lines:
                    if l.find(')') != -1:
                        tmp = tmp + l 
                    else:
                        tmp = tmp + l + '||'
                tmp_lines.clear()
                if tmp[-3:] == '||)':
                    tmp = tmp[:-3] + ')'
                processed_lines.append(tmp.strip())
                tmp=""
            else:
                tmp_lines.append(clean_line)
                    
            # Replace newlines within the parentheses with '||'
            # processed_line = clean_line.replace('\n', '||')
            # processed_lines.append(processed_line)
        # return '\n'.join(processed_lines)
        return (processed_lines)

    def read_pdf(self):        
        if len(self.pdf_file)  == 0:
            log_message('Error: No PDF file specified.',True)
            return False
        if not is_file_available(self.pdf_file):
            log_message(f'Error: PDF file {self.pdf_file} not found.',True)
            return False
        Pdf_2_Txt=self.extract_text_and_coordinates(self.pdf_file)
        # Convert the content
        #self.__converted_content = self.convert_content(Pdf_2_Txt)
        #self.write_txt(self.__converted_content,'output.txt')
        # self.first_page = self.convert_content(self.__pages_out[0])
        self.write_txt(Pdf_2_Txt,'output.txt')
        self.first_page = self.__pages_out[0]
        #self.collect_param_values()
        return True
        
    def write_txt(self,content,file_name):
        # Write the converted content to a new text file
        with open(file_name, 'w', encoding='utf-8') as file:
            file.write(content)

    def readExcel(self,excel_file):
        # Read the excel file and store data in a dictionary with key as Account Number & Value as DataFrame
        self.__dfs = pd.read_excel(excel_file)
        dic_entry={}
        for row in range(len(self.__dfs)):
            for column in self.__dfs.columns:
                dic_entry[column] = self.__dfs.loc[row,column]
            acc_num = str(self.__dfs.loc[row,'Account_Number'])
            if acc_num in self.from_excel.keys():
                log_message("Error: Duplicated entry for Account_No [" + acc_num + "]",True)
            else:
                self.from_excel[acc_num] = dic_entry
            dic_entry={}



    def set_value(self,acc_num,account_details):
        self.from_excel[acc_num] = account_details
        if len(self.from_excel[acc_num]['pdf_file']) > 0:
            self.pdf_file=self.file_path + self.from_excel[acc_num]['pdf_file']
        else:
            log_message("Error: Input excel doesn't have pdf_file",True)
        self.account_number = acc_num
        self.account_category = self.from_excel[acc_num]['ACCOUNT_CATEGORY']
        self.market_value = self.from_excel[acc_num]['MKT_CODE_VALUES']
        self.language = self.from_excel[acc_num]['LANGUAGE']
        for key in self.from_excel[acc_num].keys():
            log_message(f'{key}:    {str(self.from_excel[acc_num][key])} ')
        
        
    def find_below_string_pos(self,str,page_out):
        is_found=False
        for line in page_out:
            if str in line:
                is_found = True 
                break

    def get_string_from_image(self,img):
        bounds = reader.readtext(img)
        # print(bounds)
        text_list = []
        pos= []
        for i in range(len(bounds)):
            a=bounds[i][1]            
            if a.isalpha():
                #the service name should have more than 2 character. sometimes, the OCR convert the icon into alpha character. So, this logic implemented
                if len(a) > 1: 
                    #sometimes, the service name comes into two line like "LEASED LINES"
                    #sample: 
                    #line1: [[146, 72], [220, 72], [220, 96], [146, 96]]
                    #line2: [[157, 91], [213, 91], [213, 109], [157, 109]]
                    #so, planning to concatinate if it comes under same X position boundries

                    x1,y1,x2,y2 = bounds[i][0][0] + bounds[i][0][1]
                    pos.append([x1,y1,x2,y2,a])
                    
        remove_line=[]
        for i in range(len(pos)):
            x1,y1,x2,y2,val = pos[i]
            for j in range(len(pos)):
                if i == j : 
                    continue
                a1,b1,a2,b2,val1 = pos[j]
                if x1 > a1 and x1 < a2:
                    if y1 > b1:
                        pos[j][4] = val1 + ' ' + val
                        remove_line.append(i)
                        msg = val1 + ' ' + val
                    else:
                        pos[i][4] = val + ' ' + val1
                        remove_line.append(j)
                        msg = val + ' ' + val1

        if len(remove_line) > 0:
            remove_line.reverse()
            for line_no in remove_line:
                pos.pop(line_no)
        a=[]
        for i in range(len(pos)):
            text_list.append(pos[i][4])
        
        return text_list

    def get_line_from_content(self,msg,page_content):
        for line in page_content.split('\n'):
            if line.find(msg) > -1:
                return line
            
    def get_xy_from_msg(self,msg,page_content):
        for line in page_content.split('\n'):
            if line.find(msg) > -1:
                return line.strip('()').split(',')[:2]

    def validate_excel(self):
        global no_rows_account
        # Function to check if a string can be converted to a float
        def is_float(value):
            try:
                float(value)
                return True
            except ValueError:
                return False
            
        exl_input=self.from_excel[self.account_number]
        if is_float(exl_input['Tax']):
            exl_input['Tax'] = float(exl_input['Tax'])
        else:
            log_message("Error: unable to convert Tax value into float.",True)

        # df = pd.read_excel('Language_Strings.xlsx', sheet_name='Bus-' + self.language)
        # profile_page1=[]
        # for i in range(len(df)):
        #     profile_page1.append(df.loc[i,'Values'])
        first_page=self.first_page.split('\n')

        log_message("Info: Verifiying Page 1")
        #point #2
        # Verifing the text 'Page' in first page      
        # log_message('Info: Validating text wrt Language :')
        msg=self.__lang_string['Page']
        if msg not in first_page[0]:
            log_message(f'Error: this text {msg} is missing',True)
        
        # Verifing the text 'Invoice'
        msg=self.__lang_string['INVOICE']
        # Considering if Invoice is not matching Language is different. 
        if msg not in first_page[1]:
            log_message(f'Error: this text {msg} is missing',True)
            log_message('Error: Suspecting the language is different. Exiting.',True)
            PointCompletion["Point2"] = False
            return
        
        #Point2 Concatinate 
        def Point2_Concat(line,start_line_no,times):
            #the below logic implemented due to some language, the "Account||number" not in single line. 
            #so, finding the value and making concatinate. 

            times_cnt = 1
            #get Y value of the line
            y =  line.strip("()").split(",")[1]
            for i in range(start_line_no,start_line_no+15):
                tmp = first_page[i]
                if y == tmp.strip("()").split(",")[1]:
                    continue
                else:                    
                    if times_cnt == times:
                        tmp_content = '||' + tmp.split(",")[2]
                        line = line.replace(')',tmp_content)
                        break
                    else:
                        times_cnt += 1
            return line

        #times variable is used to find out the line number if the content is splited.
        times=0
        # Verifing the text 'Account Number'
        msg=self.__lang_string['Account_number']
        line=first_page[2]        
        if '||' not in line:
            times +=1
            line = Point2_Concat(line,2,times)
        if msg not in line:
            log_message(f'Error: this text {msg} is missing',True)
            PointCompletion["Point2"] = False

        # Verifing the text 'Invoice_number'
        msg=str(self.__lang_string['Invoice_number'])
        line=first_page[3]
        if '||' not in line:  
            times +=1
            line = Point2_Concat(line,3,times)
        if msg not in line:
            log_message(f'Error: this text {msg} is missing',True)
            PointCompletion["Point2"] = False

        # Verifing the text 'Invoice_date'
        msg=self.__lang_string['Invoice_date']
        line=first_page[4]
        if '||' not in line:              
            times +=1
            line = Point2_Concat(line,4,times)
        if msg not in line:
            log_message(f'Error: this text {msg} is missing',True)
            PointCompletion["Point2"] = False

        # Verifing the text 'Usage_period'
        msg=self.__lang_string['Usage_period']
        line=first_page[5]
        if '||' not in line:              
            times +=1
            line = Point2_Concat(line,5,times)
        if msg not in line:
            log_message(f'Error: this text {msg} is missing',True)
            PointCompletion["Point2"] = False

        # Verifing the text 'Pay_before'
        msg=self.__lang_string['Pay_before']
        line=first_page[6]
        if '||' not in line:              
            times +=1
            line = Point2_Concat(line,6,times)
        if msg not in line:
            log_message(f'Error: this text {msg} is missing',True)
            PointCompletion["Point2"] = False


        
        #finding the line number 
        line_number = 6+times
        # for i in range(6,12):
        #     if first_page[i].strip("()").split(',')[2].isnumeric():
        #         line_number = i 
        #         break
        
        #Account number value
        invoice_part= {
            'Account_number' : False,
            'Bill_ref_no'    : False, 
            'Invoice_date'   : False,
            'Usage_period'   : False,
            'Due_Date'       : False
        }
        #Account_no, bill_ref_no and other details available  y>720.
        msg=str(exl_input['BILL_REF_NO'])
        if msg == 'nan':
            log_message(f'Info: No Bill_Ref_no provided. Considering this as Proforma.')
            invoice_part['Bill_ref_no'] = True   
        for i in range(10):
            line_number +=1
            line = first_page[line_number]
            y = first_page[line_number][1:].split(',')[1]
            if float(y) < 720:
                break
            msg=self.account_number
            if msg in line:
                invoice_part['Account_number'] = True
            msg=str(exl_input['BILL_REF_NO'])
            if msg != 'nan':
                if msg in line:
                    invoice_part['Bill_ref_no'] = True                   
            msg=str(exl_input['Invoice_Date'])
            if msg in line:
                invoice_part['Invoice_date'] = True
            msg=str(exl_input['Usage_Period'])
            if msg in line:
                invoice_part['Usage_period'] = True
            msg=str(exl_input['Due_Date'])
            if msg in line:
                invoice_part['Due_Date'] = True
            
        for field in invoice_part.keys():
            if invoice_part[field] == False:
                log_message(f'Error: {field} {msg} is missing',True)
                PointCompletion["Point2"] = False



    #     line = first_page[line_number]
    #     msg=self.account_number
    #     if msg not in line:
    #         log_message(f'Error: Account_no {msg} is missing',True)
    #         PointCompletion["Point2"] = False
    #     line_number +=1
       
    #    #BILL_REF_NO value
    #     msg=str(exl_input['BILL_REF_NO'])
    #     if msg != 'nan':
    #         line = first_page[line_number]
    #         if msg not in line:
    #             log_message(f'Error: Invoice_number {msg} is missing',True)
    #             PointCompletion["Point2"] = False
    #         line_number +=1
    #     else:
    #         log_message(f'Info: No Bill_Ref_no provided. Considering this as Proforma.')

    #     #Invoice_Date value
    #     msg=exl_input['Invoice_Date']
    #     line = first_page[line_number]
    #     if msg not in line:
    #         log_message(f'Error: Invoice_date {msg} is missing',True)
    #         PointCompletion["Point2"] = False
    #     line_number +=1

    #     #Usage_Period value
    #     msg=exl_input['Usage_Period']
    #     line = first_page[line_number]
    #     if msg not in line:
    #         log_message(f'Error: Usage_period {msg} is missing',True)
    #         PointCompletion["Point2"] = False
        
    #     # Sometimes, usage period and due date will comes in single line. 
    #     # Regular expression to match duration and date
    #     pattern = r"^\d{1,2} [A-Za-z]{3} - \d{1,2} [A-Za-z]{3} \d{1,2} [A-Za-z]{3} \d{2}$"
    #     # Check if the line matches the pattern
    #     if not re.match(pattern, line.split(",", 2)[2].strip("()")):
    #         line_number +=1           
    #     #Due_Date value
    #     msg=exl_input['Due_Date']
    #     line = first_page[line_number]
    #     if msg not in line:
    #         log_message(f'Error: Due_Date {msg} is missing',True)
    #         PointCompletion["Point2"] = False
    #     line_number +=1

        #Point 3: LLA entity
        # log_message('Info: Validating LLA name :')
        mkt_code = str(exl_input['MKT_CODE'])
        total_line_no=0
        lla_line_no =0
        lla_address_line_no = 0
        owning_cost_ctr = str(exl_input['owning_cost_ctr'])
        
        for line_no in range(9,22):
            msg=self.__lang_string['Total_A']
            if msg in first_page[line_no]:
                total_line_no=line_no   

            if lla_name[mkt_code].lower() in first_page[line_no].lower():            
                lla_line_no = line_no

            if get_lla_address(mkt_code,owning_cost_ctr).lower() in first_page[line_no].lower():
                lla_address_line_no = line_no

                
        if lla_line_no == 0:
            log_message(f'Error: LLA name for market code {lla_name[mkt_code]} is missing',True)
            PointCompletion["Point3"] = False 
        
        if lla_address_line_no == 0:
            log_message(f'Error: LLA address for market code {get_lla_address(mkt_code,owning_cost_ctr)} is missing',True)
            PointCompletion["Point3"] = False 
            
        #Point 4: Total Balance Amt
        total_account_amount_a=0
        msg=self.__lang_string['Total_A']
        if total_line_no == 0:
            log_message(f'Error: this text {msg} is missing',True)
        else:
            #collecting x,y value to get the amt of total balance
            y = first_page[total_line_no][1:].split(',')[1]
            y1 = first_page[total_line_no+1][1:].split(',')[1]
            if y == y1:
                msg=first_page[total_line_no+1].split(',', 2)[-1].strip(')')
                #collect currency value 
                currency = self.from_excel[self.account_number]['CURRENCY_CODE_VALUES']
                if currency not in msg:
                    log_message(f'Error: Currency {currency} is missing',True)

                # Check if extracted_part is a float
                total_account_amount_a = msg.replace(currency,'').strip(' ').replace(',','')
                if is_float(total_account_amount_a):
                    total_account_amount_a = float(total_account_amount_a)
                else:
                    log_message(f'Error: Total Balance Amt {total_account_amount_a} is not in right format',True)
            else:
                log_message('Error: Total account amount [A] is missing',True)

        def check_next_line(total_line_no,currency):
            if currency in first_page[total_line_no+1].replace('  ',' '):
                return total_line_no+1
            else:
                log_message(f'Error: this text {currency} is missing',True)
                return total_line_no

        # Verifing the text 'My Current Charges' in first page        
        msg=self.__lang_string['My_current'] 
        if msg.replace('  ',' ') not in first_page[total_line_no+2].replace('  ',' '):
            log_message(f'Error: this text {msg} is missing',True)
        if currency not in first_page[total_line_no+2].replace('  ',' '):
            total_line_no=check_next_line(total_line_no+2, currency)
        else:
            total_line_no =total_line_no+2
        
        # Verifing the text 'Account Summary' in first page
        msg=self.__lang_string['Account_summary'] 
        if msg.replace('  ',' ') not in first_page[total_line_no+1].replace('  ',' '):
            log_message(f'Error: this text {msg} is missing',True)
        if currency not in first_page[total_line_no+1].replace('  ',' '):
            total_line_no=check_next_line(total_line_no+1, currency)
        else:
            total_line_no =total_line_no+1

        # Verifing the text 'Previous Balance' in first page        
        msg=self.__lang_string['Previous_bal']
        if msg.replace('  ',' ') not in first_page[total_line_no+1].replace('  ',' '):
            log_message(f'Error: this text {msg} is missing',True)
        else:
            total_line_no =total_line_no+1
        # Verifing  'Previous Balance Amount' in first page 
        Payments_Received=0
        Previous_Balance=0
        msg=first_page[total_line_no+1].split(',', 2)[-1].strip(')').replace(',','')
        if msg.find('||') == -1:
            log_message(f'Error: improper Previous Balance Amount {msg}',True)
        else:
            if is_float(msg.split('||')[0]):
                Previous_Balance = float(msg.split('||')[0])
            else:
                log_message(f'Error: Previous Balance Amount {msg} is not in right format',True)
            if is_float(msg.split('||')[1]):
                Payments_Received = float(msg.split('||')[1])
            else:
                log_message(f'Error: Payments Received {msg} is not in right format',True)
            total_line_no =total_line_no+1

        # Verifing the text 'Remaining Balance' in first page        
        msg=self.__lang_string['Remaining_bal']
        remaining_bal_line = total_line_no+1
        if msg.replace('  ',' ') not in first_page[total_line_no+1].replace('  ',' '):
            log_message(f'Error: this text {msg} is missing',True)

        # Verifing  'Remaining Balance Amount' in first page 
        Remaining_Balance=Current_Charges=0
        msg=first_page[remaining_bal_line+1].split(',', 2)[-1].strip(')').replace(',','')
        if msg.find('||') == -1:
            log_message(f'Error: improper Remaining Balance Amount {msg}',True)
        else:
            if is_float(msg.split('||')[0]):
                Remaining_Balance = float(msg.split('||')[0])
            else:
                log_message(f'Error: Remaining Balance Amount {msg} is not in right format',True)
            if is_float(msg.split('||')[0]):
                Current_Charges = float(msg.split('||')[1])
            else:
                log_message(f'Error: Current Charges {msg} is not in right format',True)
        

        #Point 6:
        #get "my current charges" image from pdf
        #use Harsha method to download 2
        #collecting service charge list 
        service_key_list=[]
        service_list={}
        img= self.account_number + '_page_1_image_3.png'
        img = os.path.join("extracted_images",self.account_number + '_page_1_image_3.png')
        a = self.get_string_from_image(img)
        img = os.path.join("extracted_images",self.account_number + '_page_1_image_2.png')
        b = self.get_string_from_image(img)
        service_key_list = [val.lower() for val in a + b]
        tmp_value_list = []

        if len(service_key_list) == 0:
            log_message('Error: unable to read "My Current Charge" from image.',True)
        else:
            for line_no in range(2,7):
                value=first_page[remaining_bal_line+line_no].split(',', 2)[-1].strip(')').replace(',','')
                if is_float(value):
                    tmp_value_list.append(value)
                else:
                    log_message(f'Error: improper service Amount {value}',True)

            # Verifing the text 'Total Account Balance [B]' in first page  
            total_account_amount_b=0
            tax_amt_b=0
            msg=self.__lang_string['Total_B']
            Total_b_line = remaining_bal_line+7
            tax_value=exl_input['Tax']
            if msg not in first_page[Total_b_line]:
                #some times, the total_b value comes with 0. in that case tax will not be there. 
                #so, skipping the line here 
                # log_message(f'Error: this text {msg} is missing',True)
                # PointCompletion["Point5"] = False
                msg = msg.split('||')[0]
                if msg in first_page[Total_b_line] :
                    #collecting x,y value to get the amt of total balance
                    y = first_page[Total_b_line][1:].split(',')[1]
                    y1 = first_page[Total_b_line+1][1:].split(',')[1]
                    if y == y1:
                        value=first_page[Total_b_line+1].split(',',2)[-1].split('||')[0].replace(',','').replace(')','')
                        if is_float(value):
                            total_account_amount_b = float(value)
                            if total_account_amount_b !=0:
                                log_message(f'Error: this text {msg} is missing',True)
                                PointCompletion["Point5"] = False
                        else:
                            log_message(f'Error: improper Total Account Balance [B] Amount {value}',True)
                            log_message(f'Error: this text {msg} is missing',True)
                            PointCompletion["Point5"] = False
            else:
                if '%' not in first_page[Total_b_line]:
                    log_message('Error: this symbol "%" is missing near Tax',True)
                    PointCompletion["Point5"] = False 
                #collecting x,y value to get the amt of total balance
                y = first_page[Total_b_line][1:].split(',')[1]
                y1 = first_page[Total_b_line+1][1:].split(',')[1]
                if y == y1:
                    msg=first_page[Total_b_line+1].split(',', 2)[-1]
                    if msg.find('||') == -1:
                        log_message(f'Error: improper Total Account Balance [B] Amount {msg}',True)
                    else:
                        value=msg.split('||')[0].replace(',','')
                        if is_float(value):
                            total_account_amount_b = float(value)
                        else:
                            log_message(f'Error: improper Total Account Balance [B] Amount {value}',True)
                        value=msg.split('||')[1].replace(',','').strip(')')
                        if is_float(value):
                            tax_amt_b = float(value)
                            if total_account_amount_b < 0:
                                expected_tax_amt = 0
                            else:
                                expected_tax_amt= round((total_account_amount_b * exl_input['Tax'])/100,2)
                            #Tax validation need to be check at last. Since this tax needs to validate with the tax at end of the page
                            # if expected_tax_amt != tax_amt_b:
                            #     print(f"Error: Expected tax({tax_value}) for amount {total_account_amount_b} is {expected_tax_amt}. But PDF has {tax_amt_b}")
                        else:
                            log_message(f'Error: improper Tax Amount {value}',True)
                else:
                    log_message('Error: Total account amount [B] is missing',True)

            if total_account_amount_a != total_account_amount_b : 
                log_message(f'Error: Total Account Amount not matching A({str(total_account_amount_a)}) B({str(total_account_amount_b)})',True)
                PointCompletion["Point4"] = False 
           
            #Service amount collecting for 2nd line
            msg=self.__lang_string['Ref_num_msg']
            for line_no in range(2,7):
                if msg not in first_page[Total_b_line+line_no]:
                    value=first_page[Total_b_line+line_no].split(',', 2)[-1].strip(')').replace(',','')
                    if is_float(value):
                        tmp_value_list.append(value)
                    else:
                        log_message(f'Error: improper [2nd line] service Amount {value}',True)
                else:
                    break
            
            #Setting the amount for the appropriate service
            if len(tmp_value_list) != len(service_key_list):
                log_message("Error: My Current Charge's image and value not matcing",True)
            else:
                for i in range(len(service_key_list)):
                    service_list[service_key_list[i]] = float(tmp_value_list[i].replace(',', ''))

        if len(service_list.keys()) == 0:
            log_message('Problem occured while processing "My Current Charge" (after image read)',True)

        #Verifing 2nd Page onwards
        page_number= 1
        page_out= self.__pages_out[page_number].split('\n')
        log_message("Info: Verifiying Page 2")

        def verify_header(page_number,page_out):
            # Verifing the text 'Page'        
            msg=self.__lang_string['Page']
            if msg not in page_out[0]:
                log_message(f'Error: this text {msg} is missing',True)

            # Verifing the text 'Account_number'     
            msg=self.__lang_string['Account_number2']
            if msg not in page_out[1]:
                log_message(f'Error: this text {msg} is missing',True)
                PointCompletion["Point9"] = False 

            # Verifing the value 'Account_number'
            msg=str(exl_input['BILL_REF_NO']) 
            if msg == 'nan':  #In case, if its proforma
                msg=page_out[2].split(',', 2)[-1].replace(')','')
                values = msg.split('||')
                if len(values) == 2:
                                                                                                          
                                                  
                                  
                    msg=self.account_number
                    if msg not in values[0]: 
                                                                                                            
                                                      
                                             
                                         
                                                                                                          
                                                      
                                                 
                                
                                             
                        log_message(f'Error: account_number {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
                    msg=exl_input['Invoice_Date']
                    if msg not in values[1]: 
                        log_message(f'Error: Invoice_date {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False
                msg=page_out[3].split(',', 2)[-1].replace(')','')
                values = msg.split('||')
                if len(values) == 2:
                    msg=exl_input['Usage_Period']
                    if msg not in values[0]:
                        log_message(f'Error: Usage_period {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
                    msg=exl_input['Due_Date']
                    if msg not in values[1]: 
                        log_message(f'Error: Pay_before {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
            else: 
                #when its not proforma
                msg=page_out[2].split(',', 2)[-1].replace(')','')
                values = msg.split('||')
                if len(values) < 5:                    
                    log_message(f'Error: this text {msg} is missing in page number {str(page_number+1)}',True)
                    PointCompletion["Point9"] = False 
                elif len(values) == 5:
                    msg=self.account_number
                    if msg not in values[0]: 
                        log_message(f'Error: account_number {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
                    msg=exl_input['Invoice_Date']
                    if msg not in values[1]: 
                        log_message(f'Error: Invoice_date {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
                    msg=str(exl_input['BILL_REF_NO'])
                    if msg != 'nan':
                        if msg not in values[2]: 
                            log_message(f'Error: Invoice_number {msg} is missing in page {str(page_number+1)}',True)
                            PointCompletion["Point9"] = False 
                    msg=exl_input['Usage_Period']
                    if msg not in values[3]:
                        log_message(f'Error: Usage_period {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 
                    msg=exl_input['Due_Date']
                    if msg not in values[4]: 
                        log_message(f'Error: Pay_before {msg} is missing in page {str(page_number+1)}',True)
                        PointCompletion["Point9"] = False 

        #verify header
        verify_header(page_number, page_out)
        Payments_Received_b = 0

        # Verifing the text 'My Payment Details'
        Payments_Received_b_line_no = 0
        if Payments_Received != 0:
            msg=self.__lang_string['my_pay']
            for line_no in range(2,9): # checking within next 6 lines                 
                if Payments_Received_b_line_no > 0 and '||' in page_out[line_no]: 
                    # Payments_Received_b = float(page_out[line_no].split('||') [-1].replace(')','').replace(',',''))
                    value = page_out[line_no].split(',', 2)[-1].replace(')','').replace(',','').split('||')[-1]
                    if is_float(value):
                        Payments_Received_b = float(value)
                        break
                if msg in page_out[line_no]:
                    Payments_Received_b_line_no =line_no
            if Payments_Received_b_line_no == 0:
                log_message(f'Error: this text {msg} is missing',True)
                PointCompletion["Point10"] = False
            if Payments_Received != Payments_Received_b:
                log_message(f'Error: Payments_Received {Payments_Received} is not equal to Payments  {Payments_Received_b} in 2nd page',True)
                PointCompletion["Point10"] = False
        else:
            log_message('Info: Payment Received is 0. So, not checking on 2nd page.')

        # find My Charges's line no
        my_charge_line_no=0
        msg=self.__lang_string['my_charges']
        for line_no in range(3,13): # checking within next 6 lines
            if msg in page_out[line_no]:
                my_charge_line_no = line_no
                break 
        if my_charge_line_no == 0:
            log_message(f'Error: this text {msg} is missing',True)

        # Verifing the currency near 'My Charges' in 2nd page
        msg=page_out[my_charge_line_no+1]
        if currency not in msg:
            log_message(f'Error: this text {currency} is missing near "My Charges" in 2nd page',True)
            
        # Collecting MSISDN and charges
        is_pdf_completed=False
        is_grid_completed=False
        line_no = my_charge_line_no+1
        child_details={}
        is_child_tot_found=False
        is_child_msisdn_collected= False
        page_line_count= len(page_out)
        total_pages = len(self.__pages_out)
        grid_amount = []
        grid_tax = []
        grid_without_tax_amout = 0
        discription_count=0
        is_adj = False 
        is_discount = False
        is_usage = False
        is_nrc = False
        is_rc = False
        no_of_rc_amounts =0 
        total_tax_note=0
        total_charges_excl_tax	=tax_amt=total_charges_incl_tax=0 
        is_no_tax = False
        

        while not is_pdf_completed :
            line_no += 1
            if line_no > page_line_count -1 : #line_no reached more than end of line
                page_number += 1                
                if page_number >total_pages -1  : #pdf completed 
                    is_pdf_completed = True 
                    log_message('Error:  PDF has incomplete Data.',True)
                    continue
                page_out= self.__pages_out[page_number].split('\n')
                log_message("Info: Verifiying Page " + str(page_number+1))
                page_line_count= len(page_out)

                #verify header
                verify_header(page_number, page_out)                
                line_no = 3

            line = page_out[line_no]

            #collect msisdn & offer_name
            if not is_child_msisdn_collected and float(line.split(',')[0].strip('(')) < 50:
                if line.find('-') > 1 :
                    msg =line.split(',', 2)[-1].replace(')','').split('-')
                    mobile_number = msg[1].strip()
                    mobile_number = line.split('-')[-1].strip(' )')
                    offer_name = msg[0].strip()
                    is_child_msisdn_collected = True
                    grid_without_tax_amout =0

                #verify Account Charges (last child in the grid)
                msg=self.__lang_string['Account_charges']
                if msg in line:
                    mobile_number = 'Other'
                    offer_name = msg.strip()
                    is_child_msisdn_collected = True
                    grid_without_tax_amout =0

                #verify Account level Charges
                msg=self.__lang_string['Account_Mobile_Charges']
                if msg in line:
                    mobile_number = 'Account_Level_Mobile'
                    offer_name = msg.strip()
                    is_child_msisdn_collected = True
                    grid_without_tax_amout =0

            #Check text 'Total amount with and without tax' 
            if not is_child_tot_found:
                if 'T0.0' in line:
                    is_no_tax = True
                msg=self.__lang_string['tot_char_excl_tax']
                msg1=self.__lang_string['tot_char_incl_tax']
                if msg in line and msg1 in line:
                    if '%' not in line and '||0.0' not in page_out[line_no+1] and is_no_tax == False and is_discount == False:
                        log_message(f'Error: this synbol [%] is missing under {mobile_number} on page number {str(page_number+1)}.',True)
                        PointCompletion["Point13"] = False
                    if not is_child_msisdn_collected :
                        is_grid_completed = True
                    is_child_tot_found=True 
                if is_child_msisdn_collected:
                    #check any adjustment / discount / Usate in the grid
                    msg=self.__lang_string['Adjustment']
                    if msg in line:
                        is_adj = True
                    msg=self.__lang_string['Discount']
                    if msg in line:
                        is_discount = True
                        if line.find('||') < 0:
                            #validating the next line does not has description
                            if is_float(page_out[line_no+1].strip('()').split(',')[2]):
                                #if next line has amount then, considering that there is no description.
                                log_message(f'Error:  Discount description is missing under {mobile_number} on page number {str(page_number+1)}.',True)
                                PointCompletion["Point19"] = False
                            
                    msg=self.__lang_string['Usage']
                    if msg in line:
                        is_usage = True 
                        if line.find('||') < 0:
                            #validating the next line does not has description
                            if is_float(page_out[line_no+1].strip('()').split(',')[2]):
                                #if next line has amount then, considering that there is no description.
                                log_message(f'Error:  Usage description is missing under {mobile_number} on page number {str(page_number+1)}.',True)
                                PointCompletion["Point19"] = False
                    msg=self.__lang_string['NRC']
                    if msg in line:
                        is_nrc = True 
                        if line.find('||') < 0:
                            #validating the next line does not has description
                            if is_float(page_out[line_no+1].strip('()').split(',')[2]):
                                #if next line has amount then, considering that there is no description.
                                log_message(f'Error:  NRC description is missing under {mobile_number} on page number {str(page_number+1)}.',True)
                                PointCompletion["Point19"] = False
                    #validating description under Recurring / Adjustment 
                    date_pattern = r'\d+ \w+ \d{2} - \d+ \w+ \d{2}'  # Pattern to match dates
                    #description_pattern = r'\(\d+\.\d+,\d+\.\d+,[^\d\s]+\)'  # Pattern to match descriptions (text without date format)
                    description_pattern = r"([a-zA-Z\s]+)"
                    msg=self.__lang_string['Recurring']
                    msg1=self.__lang_string['Adjustment']
                    if (msg in line or msg1 in line) and float(line.split(',')[0].strip('(')) < 100:
                        # Initialize lists to hold dates and descriptions
                        dates = []
                        descriptions = []
                        line_add = 1
                        if '||' + msg1 in line: line_add = 4
                        if msg1 + '||' in line: line_add = 0
                        if msg + '||' in line: line_add = 0

                        # Check subsequent lines for dates and descriptions until the next "Recurring:" or next 5 lines
                        for j in range(line_no + line_add, line_no + line_add + 5):
                            #if it came to end of the page, 
                            if j >= len(page_out) : 
                                break

                            # Match dates
                            date_matches = re.findall(date_pattern, page_out[j])
                            if date_matches:
                                if len(date_matches) > 1:
                                    dates.extend(date_matches)
                                else:
                                    dates.extend(date_matches)

                            # Match descriptions
                            if 100 < float(page_out[j].split(',')[0].strip('(')) < 400 and not is_discount:
                                desc_matches = page_out[j].split(',')[2].strip(')').split('||')                                                               
                                # desc_matches = re.findall(description_pattern, page_out[j])
                                if len(desc_matches) > 1:
                                    descriptions.extend(desc_matches)
                                else:
                                    descriptions.extend(desc_matches)

                            if len(dates) == len(descriptions):
                                if (msg in line) and len(dates) > 0: #if recurring, note down no.of amounts 
                                    is_rc = True
                                    #in case, NRC comes first before RC in the grid, the below one will reset.
                                    #So, considering RC always be the frist charge in the grid. 
                                    no_of_rc_amounts = len(dates)
                                break

                        # Validation: Ensure each date has a corresponding description
                        if len(dates) != len(descriptions):
                            log_message(f'Error: Description is missing under {mobile_number} on page number {str(page_number+1)}.',True)
                            PointCompletion["Point19"] = False
                        continue 

                    value = line.split(',')[2:]
                    value = ','.join(value).replace(')', '').replace(',','')
                    values=[]                                 
                    if '||' in value: 
                        values = value.split('||')
                    else:
                        values.append(value)
                    for value in values:
                        #collecting amount(s) for single grid
                        if is_float(value):
                            grid_amount.append(float(value))

                            #Adding dummy value in grid_tax for Adjustment and Discounts. 
                            if (is_adj or is_discount) and no_of_rc_amounts == 0:
                                grid_tax.append(0.0)
                            elif no_of_rc_amounts > 0:
                                no_of_rc_amounts = no_of_rc_amounts -1
                                

                        #Collecting Tax(s) for single grid
                        if value[0] == 'T' and value[1] in ['0','1','2','3','4','5','6','7','8','9'] and len(value) < 6:
                            if '%' not in value:
                                log_message(f'Error: this synbol [%] is missing under {mobile_number} on page number {str(page_number)}.',True)
                                PointCompletion["Point13"] = False
                            grid_tax.append(value.replace('T','').replace('%',''))


            #Collect amount with and without tax @end of the pdf. 
            msg=self.__lang_string['tot_char_excl_tax']
            msg1=self.__lang_string['Tax']
            msg2=self.__lang_string['tot_char_incl_tax']

            if not is_child_msisdn_collected and (msg in line or msg2 in line):
                is_grid_completed = True
                #after grid completed, if tot_char_excl_tax/tot_char_incl_tax msg occurrs then, its end of the invoice. 
                if '||' in line:
                    if msg in line and msg1 in line and msg2 not in line:
                        #tot_char_excl_tax & tax only available.
                        total_tax_note = 3
                        if '%' not in line: 
                            log_message(f'Error: this synbol [%] is missing for tax at the end',True)
                            PointCompletion["Point18"] = False
                    if msg not in line and msg1 in line and msg2  in line:
                        #tax & tot_char_incl_tax only available.
                        total_tax_note = 2
                    if msg in line and msg1 in line and msg2 in line: 
                        total_tax_note = 5
                        
                else:
                    #only tot_char_excl_tax msg found and the rest of the msg in next page. 
                    #have to note down that next line will have only one amount detail
                    if msg in line: total_tax_note = 1 # only tot_char_excl_tax available
                    if msg2 in line: total_tax_note =4 # only tot_char_incl_tax available
                
            if is_grid_completed and msg not in line and msg2 not in line and total_tax_note > 0:
                if total_tax_note == 1 :
                    value=line.split(',', 2)[-1].replace(')','').replace(',','')
                    if is_float(value) : total_charges_excl_tax = float(value)
                if '||' in line:
                    values= line.split(',', 2)[-1].replace(')','').replace(',','').split('||')
                    if total_tax_note == 2:
                        if is_float(values[0]): tax_amt = float(values[0])
                        if is_float(values[1]): total_charges_incl_tax = float(values[1])
                    if total_tax_note == 3:
                        if is_float(values[0]): total_charges_excl_tax = float(values[0])
                        if is_float(values[1]): tax_amt = float(values[1])
                    if total_tax_note == 5:
                        if is_float(values[0]): total_charges_excl_tax = float(values[0])
                        if is_float(values[1]): tax_amt = float(values[1])
                        if is_float(values[2]): total_charges_incl_tax = float(values[2])
                if total_tax_note == 4:
                    value=line.split(',', 2)[-1].replace(')','').replace(',','')
                    if is_float(value): total_charges_incl_tax = float(value)
                

                    

                    
            if is_child_tot_found and '||' in line and msg not in line and not is_grid_completed:
                if len(grid_amount) != len(grid_tax):
                    log_message(f'Error: Sub amounts & tax percentage is not proper for the grid {mobile_number}',True)
                    PointCompletion["Point12"] = False

                #Calculate single grid amount 
                for i in range(len(grid_amount)):
                    grid_without_tax_amout += grid_amount[i]
                grid_without_tax_amout = round(grid_without_tax_amout,2)
                # if grid_without_tax_amout < 0:
                #     log_message(f'Error: The accumulated amount without tax is in negative value for the grid {mobile_number}',True)
                #     PointCompletion["Point12"] = False
                grid_amount = []
                grid_tax = []
                    
                values =line.split(',', 2)[-1].replace(')','').replace(',','').split('||')
                if is_float(values[0]):
                    if grid_without_tax_amout != 0:
                        if len(values) == 3:
                            total_charges_excl_tax=float(values[0])
                            tax_amt =float(values[1])
                            total_charges_incl_tax =float(values[2])
                        elif len(values) == 2:
                            total_charges_excl_tax=float(values[0])
                            total_charges_incl_tax =float(values[1])
                            #when there is discount or other reason, if the value became zero. there will not be tax. 
                            if (total_charges_excl_tax > 0 or total_charges_incl_tax > 0) and not is_no_tax :
                                log_message(f'Error: Total amount values are not proper for {mobile_number} or not tax',True)
                                PointCompletion["Point14"] = False
                        else:
                            log_message(f'Error: Total amount values are not proper for {mobile_number}',True)
                            PointCompletion["Point14"] = False
                    else:
                        if len(values) == 2:
                            total_charges_excl_tax=float(values[0])
                            total_charges_incl_tax =float(values[1])
                            tax_amt=0
                        else:
                            log_message(f'Error: Total amount values are not proper for {mobile_number}',True)
                            PointCompletion["Point14"] = False                
                    if grid_without_tax_amout != total_charges_excl_tax and not is_grid_completed:
                        log_message(f'Error: Total amount calculation not matching. expected {grid_without_tax_amout}, but, it has {total_charges_excl_tax} for {mobile_number}',True)
                        PointCompletion["Point14"] = False

            if (is_child_tot_found or is_grid_completed) and total_charges_excl_tax != 0 and total_charges_incl_tax != 0:
                #chking tax value
                tax_value = exl_input['Tax']
                expected_tax_amt = round((tax_value * total_charges_excl_tax)/100,2)
                if expected_tax_amt -0.05 < tax_amt and expected_tax_amt +0.05 < tax_amt:
                    if is_grid_completed and total_account_amount_a > 0:
                        log_message(f'Error: Tax amount is incorrect at the end',True)
                        PointCompletion["Point16"] = False
                    else:
                        log_message("Error: Tax calculation wrong for the grid " + offer_name + ' - ' + mobile_number,True)
                        PointCompletion["Point13"] = False
                    log_message(f'-> Expected tax({tax_value}) for amount {total_charges_excl_tax} is {expected_tax_amt}. But PDF has {tax_amt}',True)
                expected_incl_amt = round(total_charges_excl_tax + tax_amt,2)
                if expected_incl_amt != total_charges_incl_tax:
                    if is_grid_completed:
                        log_message('Error: Total amount Incl Tax amt is wrong at the end',True)
                        PointCompletion["Point17"] = False
                    else:
                        log_message("Error: Total amount Incl Tax amt is wrong for the grid " + offer_name + ' - ' + mobile_number,True)
                        PointCompletion["Point14"] = False
                    log_message(f'-> Expected tax incl. amount {expected_incl_amt}. But PDF has {total_charges_incl_tax}',True)
                if not is_grid_completed:
                    child_details[mobile_number] = [offer_name,total_charges_excl_tax,total_charges_incl_tax]
                else: 
                    #Consider PDF completed. 
                    is_pdf_completed = True
                    break
            
                #setting is_child_tot_found to false to collect next child details 
                is_child_tot_found = False
                is_child_msisdn_collected = False
                is_adj = False
                is_discount = False
                is_usage = False 
                is_nrc = False
                no_of_rc_amounts = 0
                total_charges_excl_tax	=tax_amt=total_charges_incl_tax=0  
                is_no_tax = False      
            # if is_grid_completed:                
            #     #Consider PDF completed. 
            #     is_pdf_completed = True
            #     break

        #Tax compare between 1st and last page. 
        if tax_amt != tax_amt_b:
            if total_account_amount_b > 0 and tax_amt_b != expected_tax_amt:
                log_message(f'Error: Expected tax({tax_value}) for amount {total_account_amount_b} is {expected_tax_amt}. But PDF has {tax_amt_b}',True)                               
            log_message(f'Error: Tax value not matching with 1st page {tax_amt_b} <> {tax_amt}',True)
            PointCompletion["Point16"] = False

        #Waiting to complete table details collected. 
        # threads['collect_acc_tbl_info'].join()
        while True:
            if self.account_number not in Proposal_text_list.keys(): 
                if self.account_number not in no_rows_account:
                    log_message(f'Info: Account number {self.account_number} : waiting to collect DB details')
                    time.sleep(5)
                else:
                    log_message(f'Error: Account number {self.account_number} not found in the table',True)
                    return False
            else:
                break
            
            
        #Collecting service details        
        original_dict = self.get_proposal_text(child_details)
        # original_dict = self.get_proposal_test(child_details)
        db_result = {k.lower(): v for k, v in original_dict.items()}
        if len(db_result.keys()) == 0:
            log_message("Error: Received empty DB result",True)
            PointCompletion["Point6"] = False
        else:
            for service in service_list.keys():
                if service_list[service]  == 0:
                    continue
                if service in [val.lower() for val in db_result.keys()]: #converting in to lower case. since some of them are not matching
                    if round(service_list[service],2) != round(db_result[service],2):
                        log_message('Error: Service amount and grid amount not matching for ' + service,True)
                        PointCompletion["Point6"] = False 

                else:
                    log_message('Error: Service [' + service + '] not found in Database',True)
                    PointCompletion["Point6"] = False 
        
        print('completed!')
        log_message('Info: Verification Completed!')
       

        
#########################################################################################
    def get_proposal_text(self,grid_detail):
        proposal_accumulate ={}
        proposal_text=''
        for external_id in grid_detail.keys():
            amount = grid_detail[external_id][2]
            if external_id.isdigit():
                for value in Proposal_text_list[self.account_number]:
                    if external_id == value['EXTERNAL_ID']:
                        proposal_text = value['PROPOSAL_TEXT']
                        break
            elif external_id == 'Account_Level_Mobile':
                proposal_text = 'MOBILE'
            elif 'DV' in external_id or 'DA0' in external_id:
                proposal_text = 'TV'
            else:
                proposal_text = external_id.lower() #when its not mobile number and its 'other'
            if proposal_text not in proposal_accumulate.keys():
                proposal_accumulate[proposal_text] = amount
            else:
                proposal_accumulate[proposal_text] = proposal_accumulate[proposal_text] + amount
        for proposal_text in proposal_accumulate.keys():
            proposal_accumulate[proposal_text] = round(proposal_accumulate[proposal_text],2)
        print(proposal_accumulate)
        return proposal_accumulate


    def get_proposal_test(self,grid_detail):
        proposal_accumulate ={}
        if not connect():
            log_message(f'Connection failed while verify service for child account', True)
            return proposal_accumulate        
        for external_id in grid_detail.keys():
            amount = grid_detail[external_id][2]
            if external_id.isdigit():
                content="""
#!/bin/bash
export ORACLE_BASE=/home/oracle/product
export ORACLE_HOME=/home/oracle/product/11.2.0/db_1
export PATH=${PATH}:$ORACLE_HOME/bin

sqlplus -s cbs_owner/comverse@cust1 << THEEND
SET SERVEROUTPUT ON FORMAT WRAPPED SIZE 1000000;
set head off;
WHENEVER SQLERROR EXIT FAILURE ROLLBACK;
WHENEVER OSERROR EXIT FAILURE ROLLBACK;

select  distinct(PROPOSAL_TEXT) from offer_values where LANGUAGE_CODE =1 and
reseller_version_id in(SELECT max(reseller_version_id) FROM reseller_version) and 
offer_id in (
select offer_id from offer_inst@main1 where offer_inst_id in (
select offer_inst_id from rc_term_inst where rc_term_inst_id in (
select rc_term_inst_id from bill_invoice_detail where bill_ref_no in (
select  max(bill_ref_no) from bill_invoice where account_no in (select account_no from external_id_equip_map_view where 
external_id='"""+ external_id +"""' and INACTIVE_DATE is null)
and prep_status=4 and prep_error_code is  null)
and type_code in (2,3)
and subscr_no > 0
and AMOUNT_POSTPAID>0)));

quit
THEEND
"""
                writeScript(content)
                log_message('Info: Checking service in DB for '+ external_id)
                output = execute('cd /tmp; chmod +x tmp.sh; ./tmp.sh | grep -i [a-z]')
                if output.find('no rows selected') > -1:
                    log_message('Error: Unable to get proposal text for external_id ' + external_id,True)
                    continue
                proposal_text = output.split('\n')[0].lower()
            elif external_id[-2] == 'DV':
                log_message('Info: Checking service in DB for '+ external_id)
                proposal_text = 'TV'
            elif external_id == 'Account_Level_Mobile':
                proposal_text = 'MOBILE'
            else:
                log_message('Info: Checking service in DB for '+ grid_detail[external_id][0])
                proposal_text = external_id.lower() #when its not mobile number and its 'other'
            log_message('Service :' + proposal_text)
            if proposal_text not in proposal_accumulate.keys():
                proposal_accumulate[proposal_text] = amount
            else:
                proposal_accumulate[proposal_text] = proposal_accumulate[proposal_text] + amount
        disconnect()
        return proposal_accumulate

def detect_overlapping_pages(filename):
    import pdfplumber
    """Detects pages with overlapping text using raw PDF object analysis."""
    overlapping_pages = set()

    with pdfplumber.open(filename) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            text_positions = {}
            char_list = page.chars  # Extract raw character objects

            if not char_list:
                continue  # Skip empty pages

            # Track text positions
            for char in char_list:
                text = char["text"]
                x_pos = round(char["x0"], 2)
                y_pos = round(char["top"], 2)
                position_key = (x_pos, y_pos)

                if position_key in text_positions:
                    overlapping_pages.add(page_num)
                    break  # Stop checking further for this page

                text_positions[position_key] = text
    if len(overlapping_pages) > 0:
        log_message(f"Overlapping text found on page number ({filename}): {overlapping_pages}",True)
        PointCompletion['Point11'] = False

def write_output_exl(result,filename):  
    # Load the workbook
    workbook = openpyxl.load_workbook(filename)

    # Get the active sheet
    sheet = workbook.active

    # Define the red color
    red_fill = openpyxl.styles.PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    # Find the last row with data
    last_row = sheet.max_row
    col_num = 0
    for value in result.split(','): 
        # Set the accumulated value in the next row and column 1
        col_num +=1
        sheet.cell(row=last_row+1, column=col_num).value = value
        if value == 'False':
            sheet.cell(row=last_row+1, column=col_num).fill = red_fill
        
    # Save the workbook
    workbook.save(filename)



def start_process(account_details,file_path,filename): #account_details is a dictionary
    global Is_success
    Is_success = True
    refresh() #Refreshing the validation result
    myPdf = handlingPDF()
    myPdf.file_path=file_path
    account_no = str(account_details['Account_Number'])
    myPdf.set_value(account_no,account_details)
    log_message("Info: Reading Language_Strings.xlsx")
    myPdf.read_xls_string('Language_Strings.xlsx')
    save_path = "extracted_images"
    page_number = 1
    image_indices = [2, 3]  # Specify the indices of the images you want to extract

    pi.extract_specific_images(myPdf.pdf_file, save_path, page_number, image_indices, myPdf.account_number)
    # Match the logo
    # print("Verifing Logo.")
    logo_result = pi.match_logo(myPdf.pdf_file, myPdf.account_number, myPdf.account_category)
    if logo_result:
        log_message("Info: Logo matched on all pages!")        
    else:
        log_message("Error: Logo not matched.",True)
        PointCompletion['Point1'] = False
        PointCompletion['Point8'] = False
    
    # Match the promo image
    promo_result = pi.match_promo(myPdf.pdf_file, myPdf.account_number, myPdf.market_value, myPdf.account_category, myPdf.language)
    if promo_result:
        log_message("Info: Promo matched!")
    else:
        log_message("Error: Promo not matched.",True)
        PointCompletion['Point7'] = False
    if myPdf.read_pdf():
        myPdf.validate_excel()
        threads['overlap_txt_verify'].join()

        import datetime
        # Get the current datetime
        current_datetime = datetime.datetime.now()

        # Format the datetime as a string
        formatted_datetime = current_datetime.strftime('%Y-%m-%d %H:%M:%S')

        result = f"{account_no},{str(account_details['BILL_REF_NO'])},{formatted_datetime},{Is_success}"

        log_message("="*100)
        log_message("Validation Report:")
        
        for key, value in PointCompletion.items():
            result += f",{value}"
            if not value:
                log_message(key + ": <<< Failed >>>")
        
        if Is_success:
            log_message(f"Success: Account_no ({account_no}) - Bill_ref_No ({str(account_details['BILL_REF_NO'])}) - success")
        else:
            log_message(f"Failed: Account_no ({account_no}) - Bill_ref_No ({str(account_details['BILL_REF_NO'])}) - Failed")
        write_output_exl(result,filename)

        log_message("="*100)
    
if __name__ == "__main__":
    log_message("Info: Reading input.xlsx")
    print("Reading Input:", datetime.now().strftime("%H:%M:%S"))
    file_path= os.getcwd() + "\\pdf\\"
    print(os.path.join(os.getcwd(), 'input.xlsx'))
    if not is_file_available(os.path.join(os.getcwd(), 'input.xlsx')):
        log_message("Error: input.xlsx not found.",True)
        exit()
    import pandas as pd
    Excel_Input=readExcel('input.xlsx') #as dict
    from pdfminer.layout import LAParams, LTTextBox
    from pdfminer.pdfpage import PDFPage
    from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
    from pdfminer.converter import PDFPageAggregator
    import paramiko
    
    
    import re 
    import threading
    client = paramiko.SSHClient()
    # external_id_list = ",".join([f"'{item}'" for item in external_id_list])
    # collect_tbl_details(external_id_list)
    thread = threading.Thread(target=collect_tbl_details)
    threads['collect_acc_tbl_info'] = thread 
    thread.start()
    # collect_tbl_details()
    import easyocr
    reader = easyocr.Reader(['en'], gpu=False) # need to run this once before calling readtext()
    # thread.join()
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")  # Format: YYYYMMDD_HHMMSS

    # Create filename with timestamp for excel file
    filename = f"output_{timestamp}.xlsx"
    # Load the workbook
    workbook = openpyxl.Workbook()
    
    # Get the active sheet
    sheet = workbook.active

    # Define the headers (split by '|')
    headers = "Account_no|Bill_ref_no|Verified time|Is Success?|Logo|Language text|LLA Entity|Total Bal match|Tax with %|Service Icon Value match|Promo box|Next page logo|Next page language text|Payment if any|Format|Total amt excl tax in grid|Tax amt in grid|Total amt incl tax in grid|Final Total amt excl tax|Final tax amt|Final Total amt incl tax|Tax with %|Description"
    header_list = headers.split('|')

    # Add headers to the first row
    sheet.append(header_list)

    # Save the workbook
    workbook.save(filename)
    
    for acc_num in Excel_Input.keys():
        pdf_file_with_path = file_path + Excel_Input[acc_num]['pdf_file']
        thread = threading.Thread(target=detect_overlapping_pages, args=(pdf_file_with_path,))
        threads['overlap_txt_verify'] = thread
        thread.start()
        # overlap_result=detect_overlapping_pages(pdf_file_with_path)
        start_process(Excel_Input[acc_num],file_path,filename)
    print("End Time  :", datetime.now().strftime("%H:%M:%S"))
        

